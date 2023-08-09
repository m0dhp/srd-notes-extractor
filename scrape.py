from openpyxl import load_workbook


################################################################

PROGRAM_STRING = 'SRD Scraper'
VERSION_STRING = 'Version 0.1'

################################################################


def get_notes():
    # Assumes each note in column 1 of "Notes" tab starts with a row with format "Note <nnn>"
    try:
        fn = "srd.xlsx"
        wb = load_workbook(fn, read_only=True, data_only=True, keep_vba=True)
        ws = wb["Notes"]
        notes = []
        for row in ws.iter_rows(min_row=0, max_col=1):
            raw_note = row[0].value
            if raw_note and raw_note.startswith("Note "):
                notes.append(f'- [ ] {raw_note[5:]}')
        wb.close()
        return(notes)
    except:
        raise
        quit()


def get_references_set():
    # Assumes "Remarks" column on "Routes" tab has format "Notes: nnn" or "Notes: <nnn> - <nnn> <etc>"
    # SRD spreadsheet must be saved in ".xlsx" file format
    try:
        fn = "srd.xlsx"
        wb = load_workbook(fn, read_only=True, data_only=True, keep_vba=True)
        ws = wb["Routes"]
        references = set()
        for row in ws.iter_rows(min_row=0, max_col=8):
            raw_remarks = row[7].value
            if raw_remarks and raw_remarks.startswith("Notes: "):
                clean_references = raw_remarks[7:].replace("-","")
                individual_references = [int(ref.strip()) for ref in clean_references.split()]
                references.update(individual_references)
        wb.close()
        return(references)
    except:
        raise
        quit()


if __name__ == '__main__':

    # print(*get_notes(), sep='\n')
    print(*sorted(list(get_references_set())), sep='\n')