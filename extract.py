from openpyxl import load_workbook

################################################################

SRD_FILE_NAME = 'srd.xlsx'
DEBUG = False


################################################################


def get_notes(wb, markdown=False):
    # Assumes each note in column 1 of "Notes" tab starts with a row with format "Note <nnn>"
    # Markdown argument adds GitHub tick box markdown to start of each note number
    ws = wb['Notes']
    notes = []
    for row in ws.iter_rows(min_row=0, max_col=1):
        note_text = row[0].value
        try:
            if note_text.startswith('Note '):
                notes.append(f'{"- [ ] " if markdown else ""}{note_text[5:]}')
        except AttributeError:
            pass
    return notes


def get_references(wb):
    # Assumes "Remarks" column on "Routes" tab has format "Notes: nnn" or "Notes: <nnn> - <nnn> <etc>"
    # SRD spreadsheet must be saved in ".xlsx" file format
    ws = wb['Routes']
    references = set()
    for row in ws.iter_rows(min_row=0, max_col=8):
        remarks = row[7].value
        try:
            if remarks.startswith('Notes: '):
                clean_remarks = remarks[7:].replace('-', '')
                references.update([int(ref) for ref in clean_remarks.split()])
        except AttributeError:
            pass
    return references


def write_list(fn, l):
    with open(fn, 'w') as file:
        file.writelines(str(line) + '\n' for line in l)


if __name__ == '__main__':

    file_stem = SRD_FILE_NAME.split('.')[0]

    wb = load_workbook(SRD_FILE_NAME, read_only=True, data_only=True, keep_vba=True)

    if DEBUG:
        print('\nNOTES')
        print(*get_notes(wb), sep='\n')
        print('\nNOTES MARKDOWN')
        print(*get_notes(wb, True), sep='\n')
        print('\nREFERENCES')
        print(*sorted(list(get_references(wb))), sep='\n')
    else:
        write_list(f'notes-{file_stem}.txt', get_notes(wb))
        write_list(f'notes-md-{file_stem}.txt', get_notes(wb, True))
        write_list(f'references-{file_stem}.txt', sorted(list(get_references(wb))))

    wb.close()
